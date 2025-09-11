import openpyxl
import os

class ExcelManager:
    def __init__(self, archivo="data/registro.xlsx"):
        self.archivo = archivo
        if not os.path.exists("data"):
            os.makedirs("data")
        if not os.path.exists(self.archivo):
            self._crear_archivo()

    def _crear_archivo(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Crear hoja de Estudiantes
        ws_est = wb.create_sheet("Estudiantes")
        ws_est.append(["Id", "Legajo", "Nombre", "Apellido"])
        
        # Crear hoja de Materias
        ws_mat = wb.create_sheet("Materias")
        ws_mat.append(["Id", "Codigo", "Nombre"])
        
        # Crear hoja de Inscripciones con el nuevo campo Fecha
        ws_ins = wb.create_sheet("Inscripciones")
        ws_ins.append(["Id", "EstudianteId", "MateriaId", "Nota", "Fecha"])
        
        wb.save(self.archivo)

    def _abrir(self):
        return openpyxl.load_workbook(self.archivo)
        
    def _guardar(self, workbook):
        workbook.save(self.archivo)

    # === ESTUDIANTES ===
    def agregar_estudiante(self, data):
        wb = self._abrir()
        ws = wb["Estudiantes"]
        ws.append(data)
        wb.save(self.archivo)

    def obtener_estudiantes(self):
        wb = self._abrir()
        ws = wb["Estudiantes"]
        datos = []
        for row in ws.iter_rows(values_only=True):
            if row:
                datos.append({"Id": row[0], "Nombre": row[1], "Apellido": row[2], "Legajo": row[3]})
        return datos

    def modificar_estudiante(self, id_est, nombre, apellido, legajo):
        wb = self._abrir()
        ws = wb["Estudiantes"]
        for row in ws.iter_rows():
            if row[0].value == id_est:
                row[1].value = nombre
                row[2].value = apellido
                row[3].value = legajo
        self._guardar(wb)

    def eliminar_estudiante(self, id_est):
        wb = self._abrir()
        ws = wb["Estudiantes"]
        for row in ws.iter_rows():
            if row[0].value == id_est:
                ws.delete_rows(row[0].row)
        self._guardar(wb)

    # === MATERIAS ===
    def agregar_materia(self, datos):
        wb = self._abrir()
        ws = wb["Materias"]
        
        # Verificar si ya existe una materia con el mismo código
        codigo = datos[1]  # El código está en la segunda posición
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[1]).lower() == codigo.lower():
                raise ValueError(f"Ya existe una materia con el código: {codigo}")
                
        ws.append(datos)
        self._guardar(wb)

    def obtener_materias(self):
        wb = self._abrir()
        ws = wb["Materias"]
        datos = []
        for row in ws.iter_rows(values_only=True):
            if row:
                datos.append({
                    "Id": row[0],
                    "Codigo": row[1],
                    "Nombre": row[2]
                })
        return datos

    def obtener_materia_por_codigo(self, codigo):
        wb = self._abrir()
        ws = wb["Materias"]
        for row in ws.iter_rows(values_only=True):
            if row and str(row[1]).lower() == codigo.lower():
                return {
                    "Id": row[0],
                    "Codigo": row[1],
                    "Nombre": row[2]
                }
        return None

    def codigo_materia_existe(self, codigo):
        wb = self._abrir()
        ws = wb["Materias"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[1]).lower() == codigo.lower():
                return True
        return False

    def modificar_materia(self, codigo, nuevo_nombre):
        wb = self._abrir()
        ws = wb["Materias"]
        for row in ws.iter_rows(min_row=2):
            if str(row[1].value).lower() == codigo.lower():  # Buscar por código
                row[2].value = nuevo_nombre  # Actualizar nombre
                self._guardar(wb)
                return True
        return False

    def eliminar_materia(self, codigo):
        wb = self._abrir()
        ws = wb["Materias"]
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[1].value and str(row[1].value).strip().lower() == codigo.lower():
                ws.delete_rows(idx)
                self._guardar(wb)
                return True
        return False

    # === INSCRIPCIONES ===
    def agregar_inscripcion(self, data):
        wb = self._abrir()
        ws = wb["Inscripciones"]
        
        # Ensure the data has all required fields
        if len(data) < 4:  # At least ID, EstudianteId, MateriaId, Nota are required
            raise ValueError("Datos de inscripción incompletos")
            
        # If date is not provided, use current datetime
        if len(data) == 4:  # If Fecha is not in the data
            from datetime import datetime
            data.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
        ws.append(data)
        self._guardar(wb)

    def obtener_inscripciones(self):
        wb = self._abrir()
        ws = wb["Inscripciones"]
        datos = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row:  # Check if row is not empty
                inscripcion = {
                    "Id": row[0],
                    "EstudianteId": row[1],
                    "MateriaId": row[2],
                    "Nota": float(row[3]) if row[3] is not None and str(row[3]).strip() else None
                }
                # Add Fecha if it exists in the row
                if len(row) > 4:
                    inscripcion["Fecha"] = row[4]
                datos.append(inscripcion)
        return datos

    def calcular_promedio(self, estudiante_id):
        inscripciones = self.obtener_inscripciones()
        notas = [ins["Nota"] for ins in inscripciones if ins["EstudianteId"] == estudiante_id]
        return sum(notas) / len(notas) if notas else 0.0
